import frappe
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment


def copy_styles(source_cell, target_cell):
    """
    Copies the style of a source cell to a target cell.
    """
    target_cell.font = source_cell.font
    target_cell.border = source_cell.border
    target_cell.fill = source_cell.fill
    target_cell.number_format = source_cell.number_format
    target_cell.protection = source_cell.protection
    target_cell.alignment = source_cell.alignment


def create_cable_schedule_excel(revision_id):
    """
    Creates an Excel sheet for the cable schedule based on the specified revision ID.
    """
    cable_schedule_revision = frappe.get_doc(
        "Cable Schedule Revisions", revision_id
    ).as_dict()
    excel_payload = cable_schedule_revision.get("excel_payload")
    cable_schedule_data = frappe.parse_json(excel_payload)

    template_path = frappe.frappe.get_app_path(
        "thermax_backend", "templates", "heating_cable_schedule_template.xlsx"
    )

    template_workbook = load_workbook(template_path)
    cable_schedule_sheet = template_workbook["Cable Schedule_C0"]

    template_row_number = 11
    current_row = template_row_number
    template_range_start_col = 1  # Column A
    template_range_end_col = 18  # Column R (R is the 18th column)

    for main_index, cable_schedule_index in enumerate(cable_schedule_data, start=1):
        cable_schedule = cable_schedule_data[cable_schedule_index]
        # Append the main index number in the first column
        for col in range(template_range_start_col, template_range_end_col + 1):
            template_cell = cable_schedule_sheet.cell(
                row=template_row_number, column=col
            )
            target_cell = cable_schedule_sheet.cell(row=current_row, column=col)
            target_cell._style = template_cell._style

        cable_schedule_sheet.cell(
            row=current_row, column=template_range_start_col
        ).value = str(main_index)
        cable_schedule_sheet.merge_cells(f"B{current_row}:R{current_row}")
        cable_schedule_sheet["B" + str(current_row)].value = cable_schedule.get(
            "motor_name"
        )
        cable_schedule_sheet["B" + str(current_row)].alignment = Alignment(
            horizontal="left", vertical="center"
        )
        current_row += 1

        cables = cable_schedule_data[cable_schedule_index].get("cables", [])
        number_of_cables = len(cables)
        for sub_index, cable in enumerate(cables, start=1):
            # Append the sub-index number in the first column

            hierarchical_index = f"{main_index}.{sub_index}"
            for col in range(template_range_start_col, template_range_end_col + 1):
                template_cell = cable_schedule_sheet.cell(
                    row=template_row_number, column=col
                )
                target_cell = cable_schedule_sheet.cell(row=current_row, column=col)
                target_cell.font = Font(bold=False)
                target_cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

            cable_schedule_sheet.cell(row=current_row, column=1).value = (
                hierarchical_index
            )
            cable_schedule_sheet.cell(row=current_row, column=1).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            cable_schedule_sheet.cell(row=current_row, column=2).value = cable.get(
                "panel_name"
            )
            cable_schedule_sheet.cell(row=current_row, column=3).value = cable.get(
                "starter_type"
            )
            cable_schedule_sheet.cell(row=current_row, column=4).value = cable.get(
                "name"
            )
            cable_schedule_sheet.cell(row=current_row, column=5).value = cable.get(
                "voltage"
            )
            cable_schedule_sheet.cell(row=current_row, column=6).value = cable.get("kw")
            cable_schedule_sheet.cell(row=current_row, column=7).value = cable.get(
                "type_of_cable"
            )
            cable_schedule_sheet.cell(row=current_row, column=8).value = cable.get(
                "scope"
            )
            cable_schedule_sheet.cell(row=current_row, column=9).value = cable.get(
                "number_of_runs"
            )
            cable_schedule_sheet.cell(row=current_row, column=10).value = cable.get(
                "pair_core"
            )
            cable_schedule_sheet.cell(row=current_row, column=11).value = cable.get(
                "sizemm2"
            )
            cable_schedule_sheet.cell(row=current_row, column=12).value = cable.get(
                "appx_length"
            )
            cable_schedule_sheet.cell(row=current_row, column=13).value = cable.get(
                "cable_od"
            )
            cable_schedule_sheet.cell(row=current_row, column=14).value = cable.get(
                "gland_size"
            )
            cable_schedule_sheet.cell(row=current_row, column=15).value = cable.get(
                "gland_qty"
            )
            cable_schedule_sheet.cell(row=current_row, column=16).value = cable.get(
                "reducer"
            )
            cable_schedule_sheet.cell(row=current_row, column=17).value = cable.get(
                "reducer_qty"
            )
            cable_schedule_sheet.cell(row=current_row, column=18).value = cable.get(
                "comment"
            )
            current_row += 1

    return template_workbook
