import frappe
from openpyxl import load_workbook
from thermax_backend.thermax_backend.doctype.cable_schedule_revisions.cover_sheet import (
    create_cover_sheet,
)
from thermax_backend.thermax_backend.doctype.cable_schedule_revisions.other_division_cable_schedule_excel import (
    cell_styles,
)
from openpyxl.styles import Border, Side


def create_heating_excel(cable_schedule_data, project, revision_data, division_name):
    """
    Creates an Excel sheet for the cable schedule based on the specified revision ID.
    """

    template_path = frappe.frappe.get_app_path(
        "thermax_backend", "templates", "heating_cable_schedule_template.xlsx"
    )

    template_workbook = load_workbook(template_path)
    try:
        cover_sheet = template_workbook["COVER"]

        cover_sheet = create_cover_sheet(
            cover_sheet=cover_sheet,
            project_data=project,
            revision_data=revision_data,
            division_name=division_name,
        )
        cable_schedule_sheet = template_workbook["Cable Schedule"]

        template_row_number = 11
        current_row = template_row_number
        template_range_start_col = 1  # Column A
        template_range_end_col = 18  # Column R (R is the 18th column)

        for main_index, cable_schedule_index in enumerate(cable_schedule_data, start=1):
            cable_schedule = cable_schedule_data[cable_schedule_index]
            cables = cable_schedule_data[cable_schedule_index].get("cables", [])

            if not cables:  # Check for empty array
                panel_name = None  # Default value for empty cables
                starter_type = None
            else:
                # Access panel_name safely from the first element
                panel_name = cables[0].get("panel_name") if cables[0] else None
                starter_type = cables[0].get("starter_type") if cables[0] else None

            # Append the main index number in the first column
            for col in range(template_range_start_col, template_range_end_col + 1):
                template_cell = cable_schedule_sheet.cell(
                    row=template_row_number, column=col
                )
                target_cell = cable_schedule_sheet.cell(row=current_row, column=col)
                target_cell._style = template_cell._style

            cable_schedule_sheet.cell(
                row=current_row, column=template_range_start_col
            ).value = str(main_index)
            cable_schedule_sheet.merge_cells(f"B{current_row}:R{current_row}")
            cable_schedule_sheet["B" + str(current_row)].value = cable_schedule.get(
                "motor_name"
            )
            cable_schedule_sheet["B" + str(current_row)].alignment = cell_styles.get(
                "left_center"
            )
            current_row += 1

            number_of_cables = len(cables)
            cable_schedule_sheet.merge_cells(
                f"B{current_row}:B{current_row + number_of_cables - 1}"
            )
            cable_schedule_sheet.cell(current_row, 2).value = panel_name
            cable_schedule_sheet.cell(current_row, 2).alignment = cell_styles.get(
                "center"
            )
            cable_schedule_sheet.merge_cells(
                f"C{current_row}:C{current_row + number_of_cables - 1}"
            )
            cable_schedule_sheet.cell(current_row, 3).value = starter_type
            cable_schedule_sheet.cell(current_row, 3).alignment = cell_styles.get(
                "center"
            )

            for sub_index, cable in enumerate(cables, start=1):
                # Append the sub-index number in the first column

                hierarchical_index = f"{main_index}.{sub_index}"
                for col in range(template_range_start_col, template_range_end_col + 1):
                    template_cell = cable_schedule_sheet.cell(
                        row=template_row_number, column=col
                    )
                    target_cell = cable_schedule_sheet.cell(row=current_row, column=col)
                    # target_cell.font = cell_styles.get("bold")
                    target_cell.border = cell_styles.get("thin_border")

                cable_schedule_sheet.cell(row=current_row, column=1).value = (
                    hierarchical_index
                )
                cable_schedule_sheet.cell(row=current_row, column=1).alignment = (
                    cell_styles.get("center")
                )

                cable_schedule_sheet.cell(row=current_row, column=4).value = cable.get(
                    "name"
                )
                cable_schedule_sheet.cell(row=current_row, column=5).value = cable.get(
                    "voltage"
                )
                cable_schedule_sheet.cell(row=current_row, column=6).value = cable.get(
                    "kw"
                )
                cable_schedule_sheet.cell(row=current_row, column=7).value = cable.get(
                    "type_of_cable"
                )
                cable_schedule_sheet.cell(row=current_row, column=8).value = cable.get(
                    "scope"
                )
                number_of_runs = cable.get("number_of_runs", 0)
                cable_schedule_sheet.cell(row=current_row, column=9).value = int(
                    number_of_runs
                )
                cable_schedule_sheet.cell(row=current_row, column=10).value = cable.get(
                    "pair_core"
                )
                sizemm2 = cable.get("sizemm2", 0)
                cable_schedule_sheet.cell(row=current_row, column=11).value = sizemm2
                cable_schedule_sheet.cell(row=current_row, column=12).value = cable.get(
                    "appx_length"
                )
                cable_od = cable.get("cable_od", 0)
                try:
                    cable_schedule_sheet.cell(row=current_row, column=13).value = round(
                        float(cable_od), 2
                    )
                except ValueError:
                    cable_schedule_sheet.cell(row=current_row, column=13).value = 0
                # N column is for Gland Size
                cable_schedule_sheet.cell(row=current_row, column=14).value = cable.get(
                    "gland_size"
                )
                gland_qty = cable.get("gland_qty", 0)
                try:
                    cable_schedule_sheet.cell(row=current_row, column=15).value = int(
                        gland_qty
                    )
                except ValueError:
                    cable_schedule_sheet.cell(row=current_row, column=15).value = 0
                cable_schedule_sheet.cell(row=current_row, column=16).value = cable.get(
                    "reducer"
                )
                cable_schedule_sheet.cell(row=current_row, column=17).value = cable.get(
                    "reducer_qty"
                )
                cable_schedule_sheet.cell(row=current_row, column=18).value = cable.get(
                    "comment"
                )
                # cable_schedule_sheet.cell(row=current_row, column=18).style = Border(
                #     right=Side(style="thin"),
                #     left=Side(),
                #     top=Side(),
                #     bottom=Side(),
                # )
                current_row += 1

        return template_workbook
    except Exception:
        return template_workbook
