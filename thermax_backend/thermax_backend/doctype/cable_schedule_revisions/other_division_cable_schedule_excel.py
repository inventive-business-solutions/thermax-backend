import re
from collections import defaultdict

import frappe
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font

from thermax_backend.thermax_backend.doctype.cable_schedule_revisions.cover_sheet import (
    create_cover_sheet,
)
from thermax_backend.thermax_backend.doctype.cable_schedule_revisions.excel_formulae import (
    get_47_au_column_formula,
    get_48_av_column_formula,
    get_49_aw_column_formula,
    get_50_ax_column_formula,
    get_53_ba_column_formula,
    get_54_bb_column_formula,
    get_55_bc_column_formula,
    get_56_bd_column_formula,
    get_60_bh_column_formula,
    get_61_bi_column_formula,
    get_62_bj_column_formula,
    get_63_bk_column_formula,
    get_66_bn_column_formula,
    get_67_bo_column_formula,
    get_68_bp_column_formula,
    get_69_bq_column_formula,
)
from thermax_backend.thermax_backend.doctype.cable_schedule_revisions.excel_styles import (
    cell_styles,
    get_center_border_bold_style,
    get_center_border_style,
    get_left_center_style,
)


def get_yes_no_dropdown():
    """
    Get excel cell dropdown values for yes/no selection.
    """
    dropdown_values = ["Yes", "No"]
    formula = f'"{",".join(dropdown_values)}"'
    dropdown = DataValidation(type="list", formula1=formula, allow_blank=True)
    dropdown.error = "Invalid value"
    dropdown.errorTitle = "Invalid Input"
    dropdown.prompt = "Please select a value from the dropdown"
    dropdown.promptTitle = "Dropdown Input"
    return dropdown


def get_size_selection_dropdown():
    """
    Get excel cell dropdown values for cable size selection.
    """
    formula = "Dropdowns!A:A"
    dropdown = DataValidation(type="list", formula1=formula, allow_blank=True)
    dropdown.error = "Invalid value"
    dropdown.errorTitle = "Invalid Input"
    dropdown.prompt = "Please select a value from the dropdown"
    dropdown.promptTitle = "Dropdown Input"
    return dropdown


def extract_cable_type(type_of_cable):
    """
    Extracts the cable type from the type_of_cable string.
    """
    if not type_of_cable:
        return ""

    try:
        # Ensure the input is a string
        type_of_cable = str(type_of_cable)

        # Split the string by hyphen and check if there are at least two parts
        parts = type_of_cable.split("-")
        if len(parts) > 1:
            return parts[1].strip()
        else:
            return ""
    except Exception as e:
        return ""


def extract_cable_name(type_of_cable):
    """
    Extracts the cable name from the type_of_cable string.
    """
    type_of_cable = str(type_of_cable)
    if not type_of_cable:
        return ""  # Return empty string if type_of_cable is None or empty

    if "Power" in type_of_cable:
        return "PC"
    elif "Control" in type_of_cable:
        return "CC"
    elif "Signal" in type_of_cable:
        return "SC"
    else:
        return "Unknown"


def extract_number(pair_core):
    """
    Extracts the numeric part from the pair_core string.
    """
    pair_core = str(pair_core)
    if not pair_core:
        return 0  # Return 0 if pair_core is None or empty

    match = re.search(r"\d+", pair_core)  # Find the first occurrence of a number
    return int(match.group()) if match else 0  # Return the number if found, otherwise 0


def distribute_data_by_panel(rearranged_data):
    """
    Distributes the rearranged data into a dictionary organized by unique panel names.
    """
    panel_data = {}

    for item in rearranged_data:
        panel_name = item.get("panel_name")  # Extract the panel name
        if not panel_name:
            continue  # Skip items without a panel name

        if panel_name not in panel_data:
            panel_data[panel_name] = []  # Initialize a list for a new panel name

        panel_data[panel_name].append(
            item
        )  # Append the current item to the appropriate panel list

    return panel_data


def rearrange_cable_schedule_data(cable_schedule_data):
    """
    Rearranges the cable schedule data to be used in the Excel sheet.
    """
    rearranged_data = []
    for cable_schedule_index in cable_schedule_data:
        cable_schedule = cable_schedule_data[cable_schedule_index]
        cables = cable_schedule_data[cable_schedule_index].get("cables", [])
        for cable in cables:
            cable["motor_name"] = cable_schedule.get("motor_name")
            rearranged_data.append(cable)

    rearranged_data = distribute_data_by_panel(rearranged_data)

    return rearranged_data


# Function to write table headers
def write_table_headers(sheet, start_row, headers, style):
    for col, header in enumerate(headers, start=6):  # Start from column 6 (F)
        sheet.cell(start_row, col).style = style
        sheet.cell(start_row, col).value = header


# Function to write table rows
def write_table_rows(sheet, start_row, data, style, unit):
    serial_number = 1
    for key, value in data.items():
        sheet.cell(start_row, 6).style = style
        sheet.cell(start_row, 6).value = serial_number

        sheet.cell(start_row, 7).style = style
        sheet.cell(start_row, 7).value = key

        sheet.cell(start_row, 8).style = style
        sheet.cell(start_row, 8).value = value

        sheet.cell(start_row, 9).style = style
        sheet.cell(start_row, 9).value = unit

        serial_number += 1
        start_row += 1
    return start_row  # Return the next row after the table ends


def create_other_division_excel(
    cable_schedule_data, project, revision_data, division_name
):
    """
    Creates an Excel sheet for the cable schedule based on the specified revision ID.
    """
    template_path = frappe.frappe.get_app_path(
        "thermax_backend", "templates", "cable_schedule_other_division.xlsx"
    )

    template_workbook = load_workbook(template_path)

    cover_sheet = template_workbook["COVER"]
    cover_sheet = create_cover_sheet(
        cover_sheet=cover_sheet,
        project_data=project,
        revision_data=revision_data,
        division_name=division_name,
    )
    center_border_style = get_center_border_style()
    left_center_style = get_left_center_style()
    center_border_bold = get_center_border_bold_style()
    template_workbook.add_named_style(center_border_style)
    template_workbook.add_named_style(left_center_style)
    template_workbook.add_named_style(center_border_bold)

    cable_schedule_sheet = template_workbook["MCC CABLE SCHDULE"]
    # size_selection_dropdowns = get_size_selection_dropdown()
    # cable_schedule_sheet.add_data_validation(size_selection_dropdowns)

    # yes_no_dropdown = get_yes_no_dropdown()
    # cable_schedule_sheet.add_data_validation(yes_no_dropdown)

    panel_wise_data = rearrange_cable_schedule_data(cable_schedule_data)

    template_row_number = 8
    current_row = template_row_number
    power_cable_size = {}
    control_cable_size = {}
    gland_nos = {}
    shroud_nos = {}

    for panel_name, panel_cables in panel_wise_data.items():
        # Merge and style the panel name row
        cable_schedule_sheet.cell(current_row, 1).style = "left_center_style"
        cable_schedule_sheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=32,
        )
        cable_schedule_sheet.row_dimensions[current_row].height = 40

        cable_schedule_sheet.cell(current_row, 1).value = panel_name

        current_row += 1  # Move to the next row after panel name

        motor_name_groups = defaultdict(list)

        for pc in panel_cables:
            motor_name = pc.get(
                "motor_name", "Unknown"
            )  # Default to "Unknown" if motor_name is missing
            motor_name_groups[motor_name].append(pc)

        motor_name_groups = dict(motor_name_groups)

        for motor_number_index, (motor_name, motor_cables) in enumerate(
            motor_name_groups.items()
        ):
            total_panel_cores = 0
            for cable in motor_cables:
                pair_core = cable.get("pair_core")
                number_of_cores = extract_number(pair_core)
                total_panel_cores += number_of_cores

            cable_schedule_sheet.cell(current_row, 1).style = "center_border_style"
            cable_schedule_sheet.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row + max(total_panel_cores - 1, 0),
                end_column=1,
            )
            cable_schedule_sheet.cell(current_row, 1).value = motor_number_index + 1

            # 7th column G: FROM PANEL / JB NO
            cable_schedule_sheet.cell(current_row, 7).style = "center_border_style"
            cable_schedule_sheet.merge_cells(
                start_row=current_row,
                start_column=7,
                end_row=current_row + max(total_panel_cores - 1, 0),
                end_column=7,
            )
            cable_schedule_sheet.cell(current_row, 7).value = panel_name

            # 8th column H: FROM PANEL / JB DESCRIPTION
            cable_schedule_sheet.cell(current_row, 8).style = "center_border_style"
            cable_schedule_sheet.merge_cells(
                start_row=current_row,
                start_column=8,
                end_row=current_row + max(total_panel_cores - 1, 0),
                end_column=8,
            )
            cable_schedule_sheet.cell(current_row, 8).value = panel_name

            # 9th column I: SYSTEM VOLTAGE
            cable_schedule_sheet.cell(current_row, 9).style = "center_border_style"
            cable_schedule_sheet.merge_cells(
                start_row=current_row,
                start_column=9,
                end_row=current_row + max(total_panel_cores - 1, 0),
                end_column=9,
            )
            supply_voltage = (
                motor_cables[0].get("voltage") if len(motor_cables) > 1 else 0
            )
            cable_schedule_sheet.cell(current_row, 9).value = supply_voltage

            # 10th column J: KW RATING
            cable_schedule_sheet.cell(current_row, 10).style = "center_border_style"
            cable_schedule_sheet.merge_cells(
                start_row=current_row,
                start_column=10,
                end_row=current_row + max(total_panel_cores - 1, 0),
                end_column=10,
            )
            cable_schedule_sheet.cell(current_row, 10).value = (
                motor_cables[0].get("kw") if len(motor_cables) > 1 else 0
            )

            for cable_index, cable in enumerate(motor_cables, start=1):
                # Extract the number of cores from the pair_core string
                pair_core = cable.get("pair_core")
                number_of_cores = extract_number(pair_core)

                # Calculate the ending row for the merge
                end_row = current_row + max(
                    number_of_cores - 1, 0
                )  # Ensure non-negative range

                # 2nd column B: CABLE NO.
                cable_schedule_sheet.cell(current_row, 2).style = "center_border_style"
                cable_schedule_sheet.merge_cells(
                    start_row=current_row,
                    start_column=2,
                    end_row=end_row,
                    end_column=2,
                )
                cable_schedule_sheet.cell(
                    current_row, 2
                ).value = f"""{extract_cable_name(
                    cable.get("type_of_cable")
                )} - """

                # 3rd column C: CABLE TAG
                cable_schedule_sheet.cell(current_row, 3).style = "center_border_style"
                cable_schedule_sheet.merge_cells(
                    start_row=current_row,
                    start_column=3,
                    end_row=end_row,
                    end_column=3,
                )
                cable_schedule_sheet.cell(current_row, 3).value = (
                    f"""=CONCATENATE(B{current_row},"/",G{current_row},"-",P{current_row})"""
                )

                # 11th column K: TYPE
                cable_schedule_sheet.cell(current_row, 11).style = "center_border_style"
                cable_schedule_sheet.merge_cells(
                    start_row=current_row,
                    start_column=11,
                    end_row=end_row,
                    end_column=11,
                )
                cable_schedule_sheet.cell(current_row, 11).value = cable.get(
                    "starter_type"
                )

                # 12th column L: CABLE SIZE
                cable_schedule_sheet.cell(current_row, 12).style = "center_border_style"
                cable_schedule_sheet.merge_cells(
                    start_row=current_row,
                    start_column=12,
                    end_row=end_row,
                    end_column=12,
                )

                size_key = f"{cable.get('pair_core')} X {cable.get('sizemm2')} SQ.MM. {cable.get('cable_material')} {cable.get('type_of_insulation')} ARMOURED CABLE"

                cable_schedule_sheet.cell(current_row, 12).value = size_key
                cable_length = cable.get("appx_length", 0)

                if "Power" in cable.get("type_of_cable"):
                    if size_key in power_cable_size:
                        power_cable_size[size_key] += cable_length
                    else:
                        power_cable_size[size_key] = cable_length

                if "Control" in cable.get("type_of_cable"):
                    if size_key in control_cable_size:
                        control_cable_size[size_key] += cable_length
                    else:
                        control_cable_size[size_key] = cable_length

                # 13th column M: CABLE TYPE
                cable_schedule_sheet.cell(current_row, 13).style = "center_border_style"
                cable_schedule_sheet.merge_cells(
                    start_row=current_row,
                    start_column=13,
                    end_row=end_row,
                    end_column=13,
                )
                cable_schedule_sheet.cell(current_row, 13).value = extract_cable_type(
                    cable.get("type_of_cable")
                )

                # 16th column P: TAG NOS.
                cable_schedule_sheet.cell(current_row, 16).style = "center_border_style"
                cable_schedule_sheet.merge_cells(
                    start_row=current_row,
                    start_column=16,
                    end_row=end_row,
                    end_column=16,
                )
                cable_schedule_sheet.cell(current_row, 16).value = cable.get(
                    "tag_number"
                )

                # 18th column R: TO EQUIPMENT DESCRIPTION
                cable_schedule_sheet.cell(current_row, 18).style = "center_border_style"
                cable_schedule_sheet.merge_cells(
                    start_row=current_row,
                    start_column=18,
                    end_row=end_row,
                    end_column=18,
                )
                cable_schedule_sheet.cell(current_row, 18).value = cable.get(
                    "service_description"
                )

                # 19th column S: CABLE LENGTH (IN MTR.)
                cable_schedule_sheet.cell(current_row, 19).style = "center_border_style"
                cable_schedule_sheet.merge_cells(
                    start_row=current_row,
                    start_column=19,
                    end_row=end_row,
                    end_column=19,
                )
                cable_schedule_sheet.cell(current_row, 19).value = cable.get(
                    "appx_length"
                )

                # 20th column T: CABLE OD
                cable_schedule_sheet.cell(current_row, 20).style = "center_border_style"
                cable_schedule_sheet.merge_cells(
                    start_row=current_row,
                    start_column=20,
                    end_row=end_row,
                    end_column=20,
                )
                cable_schedule_sheet.cell(current_row, 20).value = cable.get("cable_od")

                # 21st column U: ENTRY AVAILABLE AT PANEL / JB
                cable_schedule_sheet.cell(current_row, 21).style = "center_border_style"
                cable_schedule_sheet.merge_cells(
                    start_row=current_row,
                    start_column=21,
                    end_row=end_row,
                    end_column=21,
                )
                cable_schedule_sheet.cell(current_row, 21).value = "Plate"

                # 22nd column V: SIZE SELECTED AT PANEL / JB
                cable_schedule_sheet.cell(current_row, 22).style = "center_border_style"
                cable_schedule_sheet.merge_cells(
                    start_row=current_row,
                    start_column=22,
                    end_row=end_row,
                    end_column=22,
                )
                # size_selection_dropdowns.add(cable_schedule_sheet[f"V{current_row}"])

                # 23rd column W: GLAND CAT NO AT PANEL / JB
                cable_schedule_sheet.cell(current_row, 23).style = "center_border_style"
                cable_schedule_sheet.merge_cells(
                    start_row=current_row,
                    start_column=23,
                    end_row=end_row,
                    end_column=23,
                )
                cable_schedule_sheet.cell(current_row, 23).value = (
                    f"""=IF('GLAND SELEC. INPUT & NOTES SHT'!$H$17="Ni PLATED BRASS",IF($AQ{current_row}="NARMOURED CABLE",$AX{current_row},IF($AQ{current_row}=" ARMOURED CABLE",IF($AT{current_row}="M",$AV{current_row},IF($AT{current_row}=" ",$AU{current_row},IF($AT{current_row}="N",$AW{current_row},"NA"))))),IF($AQ{current_row}="NARMOURED CABLE",$BK{current_row},IF($AQ{current_row}=" ARMOURED CABLE",IF($BG{current_row}="M",$BI{current_row},IF($BG{current_row}=" ",$BH{current_row},IF($BG{current_row}="N",$BJ{current_row},"NA"))))))"""
                )

                # 24th column X: SHROUD REQUIREMENT AT PANEL / JB
                cable_schedule_sheet.cell(current_row, 24).style = "center_border_style"
                cable_schedule_sheet.merge_cells(
                    start_row=current_row,
                    start_column=24,
                    end_row=end_row,
                    end_column=24,
                )
                # yes_no_dropdown.add(cable_schedule_sheet[f"V{current_row}"])

                # 25th column Y: SHROUD CAT NO  AT PANEL / JB
                cable_schedule_sheet.cell(current_row, 25).style = "center_border_style"
                cable_schedule_sheet.merge_cells(
                    start_row=current_row,
                    start_column=25,
                    end_row=end_row,
                    end_column=25,
                )
                cable_schedule_sheet.cell(current_row, 25).value = (
                    f"""=IF($X{current_row}="YES","PVC SHROUD FOR "&$W{current_row},"NA")"""
                )

                # 26th column Z: ENTRY AVAILABLE AT EQUIPMENT
                cable_schedule_sheet.cell(current_row, 26).style = "center_border_style"
                cable_schedule_sheet.merge_cells(
                    start_row=current_row,
                    start_column=26,
                    end_row=end_row,
                    end_column=26,
                )

                # 27th column AA: SIZE SELECTED AT EQUIPMENT
                cable_schedule_sheet.cell(current_row, 27).style = "center_border_style"
                cable_schedule_sheet.merge_cells(
                    start_row=current_row,
                    start_column=27,
                    end_row=end_row,
                    end_column=27,
                )
                # size_selection_dropdowns.add(cable_schedule_sheet[f"AA{current_row}"])

                # 28th column AB: GLAND CAT NO AT EQUIPMENT
                cable_schedule_sheet.cell(current_row, 28).style = "center_border_style"
                cable_schedule_sheet.merge_cells(
                    start_row=current_row,
                    start_column=28,
                    end_row=end_row,
                    end_column=28,
                )
                cable_schedule_sheet.cell(current_row, 28).value = (
                    f"""=IF('GLAND SELEC. INPUT & NOTES SHT'!$H$17="Ni PLATED BRASS",IF($AQ{current_row}="NARMOURED CABLE",$BD{current_row},IF($AQ{current_row}=" ARMOURED CABLE",IF($AZ{current_row}="M",$BB{current_row},IF($AZ{current_row}=" ",$BA{current_row},IF($AZ{current_row}="N",$BC{current_row},"NA"))))),IF($AQ{current_row}="NARMOURED CABLE",$BQ{current_row},IF($AQ{current_row}=" ARMOURED CABLE",IF($BM{current_row}="M",$BO{current_row},IF($BM{current_row}=" ",$BN{current_row},IF($BM{current_row}="N",$BP{current_row},"NA"))))))"""
                )

                # 29th column AC: SHROUD REQUIREMENT AT EQUIPMENT
                cable_schedule_sheet.cell(current_row, 29).style = "center_border_style"
                cable_schedule_sheet.merge_cells(
                    start_row=current_row,
                    start_column=29,
                    end_row=end_row,
                    end_column=29,
                )
                # yes_no_dropdown.add(cable_schedule_sheet[f"AC{current_row}"])

                # 30th column AD: SHROUD CAT NO AT EQUIPMENT
                cable_schedule_sheet.cell(current_row, 30).style = "center_border_style"
                cable_schedule_sheet.merge_cells(
                    start_row=current_row,
                    start_column=30,
                    end_row=end_row,
                    end_column=30,
                )
                cable_schedule_sheet.cell(current_row, 30).value = (
                    f"""=IF($AC{current_row}="YES","PVC SHROUD FOR "&$AB{current_row},"NA")"""
                )

                # 31st column AE: CABLE TRAY ROUTING
                cable_schedule_sheet.cell(current_row, 31).style = "center_border_style"
                cable_schedule_sheet.merge_cells(
                    start_row=current_row,
                    start_column=31,
                    end_row=end_row,
                    end_column=31,
                )

                # # 32nd column AF: REMARKS
                # cable_schedule_sheet.cell(current_row, 32).style = "center_border_style"
                # cable_schedule_sheet.merge_cells(
                #     start_row=current_row,
                #     start_column=32,
                #     end_row=end_row,
                #     end_column=32,
                # )
                # cable_schedule_sheet.cell(current_row, 32).value = cable.get("comment")

                for row in range(number_of_cores):
                    # 4th column D: FEEDER NO
                    cable_schedule_sheet.cell(current_row + row, 4).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.row_dimensions[current_row + row].height = 40

                    # 5th column E: FROM TB NO
                    cable_schedule_sheet.cell(current_row + row, 5).style = (
                        "center_border_style"
                    )

                    # 6th column F: FROM FERRULE NO
                    cable_schedule_sheet.cell(current_row + row, 6).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 6).value = (
                        f"""=CONCATENATE(O{current_row + row},"/",E{current_row + row})"""
                    )

                    # 14th column N: SCOPE
                    cable_schedule_sheet.cell(current_row + row, 14).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 14).value = (
                        f"""=CONCATENATE(E{current_row + row},"/",O{current_row + row})"""
                    )

                    # 15th column O: TO EQUIPMENT NO
                    cable_schedule_sheet.cell(current_row + row, 15).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 15).value = (
                        f"""=P{current_row + row}&" - "&Q{current_row + row}"""
                    )

                    # 17th column Q: TO TB NO.
                    cable_schedule_sheet.cell(current_row + row, 17).style = (
                        "center_border_style"
                    )

                    # 32nd column AF: REMARKS
                    cable_schedule_sheet.cell(current_row + row, 32).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 32).value = cable.get(
                        "comment"
                    )

                    # 33rd column AG: 50W
                    cable_schedule_sheet.cell(current_row + row, 33).style = (
                        "center_border_style"
                    )

                    # 34th column AH: 100W
                    cable_schedule_sheet.cell(current_row + row, 34).style = (
                        "center_border_style"
                    )

                    # 35th column AI: 50W
                    cable_schedule_sheet.cell(current_row + row, 35).style = (
                        "center_border_style"
                    )

                    # 36th column AJ: 100W
                    cable_schedule_sheet.cell(current_row + row, 36).style = (
                        "center_border_style"
                    )

                    # 43rd column AQ: TYPE OF CABLE
                    cable_schedule_sheet.cell(current_row + row, 43).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 43).value = (
                        f"""=RIGHT(L{current_row + row},15)"""
                    )

                    # 44th column AR: MAKE
                    cable_schedule_sheet.cell(current_row + row, 44).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 44).value = (
                        """='GLAND SELEC. INPUT & NOTES SHT'!$H$16"""
                    )

                    # 45th column AS:
                    cable_schedule_sheet.cell(current_row + row, 45).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 45).value = (
                        f"""=RIGHT($V{current_row + row},3)"""
                    )

                    # 46th column AT:
                    cable_schedule_sheet.cell(current_row + row, 46).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 46).value = (
                        f"""=LEFT($AS{current_row + row},1)"""
                    )

                    # 47th column AU:
                    cable_schedule_sheet.cell(current_row + row, 47).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 47).value = (
                        get_47_au_column_formula(current_row=current_row, row=row)
                    )

                    # 48th column AV:
                    cable_schedule_sheet.cell(current_row + row, 48).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 48).value = (
                        get_48_av_column_formula(current_row=current_row, row=row)
                    )

                    # 49th column AW:
                    cable_schedule_sheet.cell(current_row + row, 49).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 49).value = (
                        get_49_aw_column_formula(current_row=current_row, row=row)
                    )

                    # 50th column AX:
                    cable_schedule_sheet.cell(current_row + row, 50).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 50).value = (
                        get_50_ax_column_formula(current_row=current_row, row=row)
                    )

                    # 51st column AY:
                    cable_schedule_sheet.cell(current_row + row, 51).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 51).value = (
                        f"""=RIGHT($AA{current_row + row},3)"""
                    )

                    # 52nd column AZ:
                    cable_schedule_sheet.cell(current_row + row, 52).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 52).value = (
                        f"""=LEFT($AY{current_row + row},1)"""
                    )

                    # 53rd column BA:
                    cable_schedule_sheet.cell(current_row + row, 53).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 53).value = (
                        get_53_ba_column_formula(current_row=current_row, row=row)
                    )

                    # 54th column BB:
                    cable_schedule_sheet.cell(current_row + row, 54).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 54).value = (
                        get_54_bb_column_formula(current_row=current_row, row=row)
                    )

                    # 55th column BC:
                    cable_schedule_sheet.cell(current_row + row, 55).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 55).value = (
                        get_55_bc_column_formula(current_row=current_row, row=row)
                    )

                    # 56th column BD:
                    cable_schedule_sheet.cell(current_row + row, 56).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 56).value = (
                        get_56_bd_column_formula(current_row=current_row, row=row)
                    )

                    # 58th column BF:
                    cable_schedule_sheet.cell(current_row + row, 58).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 58).value = (
                        f"""=RIGHT($V{current_row + row},3)"""
                    )

                    # 59th column BG:
                    cable_schedule_sheet.cell(current_row + row, 59).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 59).value = (
                        f"""=LEFT($AS{current_row + row},1)"""
                    )

                    # 60th column BH:
                    cable_schedule_sheet.cell(current_row + row, 60).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 60).value = (
                        get_60_bh_column_formula(current_row=current_row, row=row)
                    )

                    # 61st column BI:
                    cable_schedule_sheet.cell(current_row + row, 61).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 61).value = (
                        get_61_bi_column_formula(current_row=current_row, row=row)
                    )

                    # 62nd column BJ:
                    cable_schedule_sheet.cell(current_row + row, 62).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 62).value = (
                        get_62_bj_column_formula(current_row=current_row, row=row)
                    )

                    # 63rd column BK:
                    cable_schedule_sheet.cell(current_row + row, 63).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 63).value = (
                        get_63_bk_column_formula(current_row=current_row, row=row)
                    )

                    # 64th column BL:
                    cable_schedule_sheet.cell(current_row + row, 64).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 64).value = (
                        f"""=RIGHT($AA{current_row + row},3)"""
                    )

                    # 65th column BM:
                    cable_schedule_sheet.cell(current_row + row, 65).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 65).value = (
                        f"""=LEFT($AY{current_row + row},1)"""
                    )

                    # 66th column BN:
                    cable_schedule_sheet.cell(current_row + row, 66).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 66).value = (
                        get_66_bn_column_formula(current_row=current_row, row=row)
                    )

                    # 67th column BO:
                    cable_schedule_sheet.cell(current_row + row, 67).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 67).value = (
                        get_67_bo_column_formula(current_row=current_row, row=row)
                    )

                    # 68th column BP:
                    cable_schedule_sheet.cell(current_row + row, 68).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 68).value = (
                        get_68_bp_column_formula(current_row=current_row, row=row)
                    )

                    # 69th column BQ:
                    cable_schedule_sheet.cell(current_row + row, 69).style = (
                        "center_border_style"
                    )
                    cable_schedule_sheet.cell(current_row + row, 69).value = (
                        get_69_bq_column_formula(current_row=current_row, row=row)
                    )

                # Update the current_row to the row after the merged cells
                current_row = end_row + 1

    cable_summary_sheet = template_workbook["CABLE SUMMARY"]

    # Initialize variables
    headers = ["Sr. No", "CABLE SIZE", "QTY", "UNIT"]
    bold_style = "center_border_bold_style"
    normal_style = "center_border_style"
    unit = "Mtrs."
    start_row = 5

    # Write the first table
    write_table_headers(cable_summary_sheet, start_row, headers, bold_style)
    start_row = write_table_rows(
        cable_summary_sheet, start_row + 1, power_cable_size, normal_style, unit
    )

    # Add a 4-row gap before the second table
    start_row += 4

    # Write the title for the second table
    cable_summary_sheet.cell(start_row, 7).style = bold_style
    cable_summary_sheet.cell(start_row, 7).value = "CONTROL CABLE"
    start_row += 1

    # Write the second table
    write_table_headers(cable_summary_sheet, start_row, headers, bold_style)
    write_table_rows(
        cable_summary_sheet, start_row + 1, control_cable_size, normal_style, unit
    )

    return template_workbook
