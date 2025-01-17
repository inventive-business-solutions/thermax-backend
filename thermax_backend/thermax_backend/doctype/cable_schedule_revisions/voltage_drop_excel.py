import frappe
from openpyxl import load_workbook
from thermax_backend.thermax_backend.doctype.cable_schedule_revisions.cover_sheet import (
    create_cover_sheet,
)


def create_voltage_drop_excel(revision_id):
    """
    Creates an Excel sheet for the voltage drop calculation based on the specified revision ID.
    """

    cable_schedule_revision = frappe.get_doc(
        "Cable Schedule Revisions", revision_id
    ).as_dict()
    project_id = cable_schedule_revision.get("project_id")
    project = frappe.get_doc("Project", project_id).as_dict()
    division_name = project.get("division")

    project_id = project.get("name")
    cable_schedule_data = cable_schedule_revision.get("cable_schedule_data")

    template_path = frappe.frappe.get_app_path(
        "thermax_backend", "templates", "voltage_drop_calculation_template.xlsx"
    )

    template_workbook = load_workbook(template_path)

    cover_sheet = template_workbook["COVER"]

    cover_sheet = create_cover_sheet(
        cover_sheet=cover_sheet,
        project_data=project,
        revision_data=cable_schedule_revision,
        division_name=division_name,
    )
    voltage_drop_calculation_sheet = template_workbook["VOLTAGE DROP CALCULATION"]

    total_rows = len(cable_schedule_data)
    template_row_number = 3
    dynamic_start_row_number = template_row_number + 1
    template_range_start_col = 1  # Column A
    template_range_end_col = 28  # Column AB (AB is the 28th column)

    for row in range(dynamic_start_row_number, template_row_number + total_rows):
        for col in range(template_range_start_col, template_range_end_col + 1):
            # Get the template cell
            template_cell = voltage_drop_calculation_sheet.cell(
                row=template_row_number, column=col
            )
            # Get the target cell
            target_cell = voltage_drop_calculation_sheet.cell(row=row, column=col)
            # Copy the style from the template cell
            target_cell._style = template_cell._style

            # Apply column width (only once per column)
            column_letter = template_cell.column_letter
            if (
                row == dynamic_start_row_number
            ):  # Apply width only on the first iteration for each column
                template_width = voltage_drop_calculation_sheet.column_dimensions[
                    column_letter
                ].width
                voltage_drop_calculation_sheet.column_dimensions[
                    column_letter
                ].width = template_width

    for index, data in enumerate(cable_schedule_data):
        row = template_row_number + index
        voltage_drop_calculation_sheet.cell(row=row, column=1, value=index + 1)
        voltage_drop_calculation_sheet.cell(
            row=row, column=2, value=data.get("tag_number")
        )
        voltage_drop_calculation_sheet.cell(
            row=row, column=3, value=data.get("service_description")
        )
        standby_kw = data.get("standby_kw")
        working_kw = data.get("working_kw")
        non_zero_kw = standby_kw if standby_kw >= 0 else working_kw
        voltage_drop_calculation_sheet.cell(row=row, column=4, value=non_zero_kw)
        voltage_drop_calculation_sheet.cell(row=row, column=5, value="3 Phase")
        starter_type = data.get("starter_type")
        voltage_drop_calculation_sheet.cell(row=row, column=6, value=starter_type)
        voltage_drop_calculation_sheet.cell(
            row=row, column=7, value=data.get("supply_voltage")
        )
        voltage_drop_calculation_sheet.cell(row=row, column=8, value=0.86)

        cos_running_cell = voltage_drop_calculation_sheet.cell(
            row=row, column=9, value=data.get("cos_running")
        )
        voltage_drop_calculation_sheet.cell(
            row=row, column=10, value=f"=SIN(ACOS({cos_running_cell.coordinate}))"
        )
        cos_starting_cell = voltage_drop_calculation_sheet.cell(
            row=row, column=11, value=data.get("cos_starting")
        )
        voltage_drop_calculation_sheet.cell(
            row=row, column=12, value=f"=SIN(ACOS({cos_starting_cell.coordinate}))"
        )
        running_current = float(data.get("motor_rated_current"))
        voltage_drop_calculation_sheet.cell(
            row=row, column=13, value=data.get("motor_rated_current")
        )
        # N(14) Column: I (START) in Amp
        starting_current = 0
        if starter_type == "DOL STARTER" or starter_type == "R-DOL":
            starting_current = running_current * 7.5
        else:
            starting_current = running_current * 3

        voltage_drop_calculation_sheet.cell(
            row=row, column=14, value=round(starting_current, 2)
        )
        voltage_drop_calculation_sheet.cell(
            row=row, column=15, value=data.get("cable_material")
        )
        voltage_drop_calculation_sheet.cell(
            row=row, column=16, value=data.get("number_of_runs")
        )
        voltage_drop_calculation_sheet.cell(
            row=row,
            column=17,
            value=f"{data.get('number_of_runs')} x {data.get('number_of_cores')} x {data.get('final_cable_size')}",
        )
        voltage_drop_calculation_sheet.cell(
            row=row, column=18, value=data.get("resistance_meter")
        )
        voltage_drop_calculation_sheet.cell(
            row=row, column=19, value=data.get("reactance_meter")
        )
        voltage_drop_calculation_sheet.cell(
            row=row, column=20, value=data.get("apex_length")
        )
        voltage_drop_calculation_sheet.cell(
            row=row, column=21, value=data.get("vd_running")
        )
        voltage_drop_calculation_sheet.cell(
            row=row, column=22, value=data.get("vd_starting")
        )
        voltage_drop_calculation_sheet.cell(
            row=row, column=23, value=data.get("percent_vd_running")
        )
        voltage_drop_calculation_sheet.cell(
            row=row, column=24, value=data.get("percent_vd_starting")
        )
        voltage_drop_calculation_sheet.cell(
            row=row, column=25, value=data.get("selected_cable_capacity_amp")
        )
        voltage_drop_calculation_sheet.cell(
            row=row, column=26, value=data.get("derating_factor")
        )
        voltage_drop_calculation_sheet.cell(
            row=row, column=27, value=data.get("final_capacity")
        )
        voltage_drop_calculation_sheet.cell(
            row=row, column=28, value=data.get("cable_selected_status")
        )
    return template_workbook
