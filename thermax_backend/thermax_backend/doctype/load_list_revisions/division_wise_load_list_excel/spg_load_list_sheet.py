import frappe
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def get_spg_load_list_excel(
    template_workbook,
    electrical_load_list_data,
    panels_data,
    incomer_power_supply,
    incomer_db_data,
):
    """
    Gets the Excel workbook for the "Heating" or "WWS SPG" division.
    """

    load_list_output_sheet = template_workbook["LOAD LIST OUTPUT"]
    all_panels_sheet = template_workbook.copy_worksheet(load_list_output_sheet)
    number_of_panels = len(panels_data)
    panel_name = "All Panels"
    if number_of_panels == 1:
        panel_name = electrical_load_list_data[0].get("panel")

    all_panels_sheet = create_spg_load_list_excel(
        electrical_load_list_data=electrical_load_list_data,
        load_list_output_sheet=all_panels_sheet,
        incomer_power_supply=incomer_power_supply,
        number_of_panels=number_of_panels,
        panel_name=panel_name,
        incomer_db_data=incomer_db_data,
    )

    if number_of_panels > 1:
        for panel_name, panel_data in panels_data.items():
            panel_sheet = template_workbook.copy_worksheet(load_list_output_sheet)
            panel_sheet.title = panel_name

            panel_sheet = create_spg_load_list_excel(
                electrical_load_list_data=panel_data,
                load_list_output_sheet=panel_sheet,
                incomer_power_supply=incomer_power_supply,
                number_of_panels=number_of_panels,
                panel_name=panel_name,
                incomer_db_data=incomer_db_data,
            )

    template_workbook.remove(load_list_output_sheet)
    all_panels_sheet.title = "LOAD LIST OUTPUT"
    return template_workbook


def create_spg_load_list_excel(
    electrical_load_list_data,
    load_list_output_sheet,
    incomer_power_supply,
    number_of_panels,
    panel_name,
    incomer_db_data,
):
    """
    Generates an Excel sheet for the electrical load list for the "Heating" or "WWS SPG" division.
    Args:
        electrical_load_list_data (list): The data representing the electrical load list.
    Returns:
        object: The updated Excel worksheet object containing the load list.
    """

    total_rows = len(electrical_load_list_data)
    template_row_number = 3
    dynamic_start_row_number = template_row_number + 1
    template_range_start_col = 1  # Column A
    template_range_end_col = 17  # Column Q (17 the column number)

    # Get the row height of the template row
    template_row_height = load_list_output_sheet.row_dimensions[
        template_row_number
    ].height

    load_list_output_sheet.insert_rows(dynamic_start_row_number, total_rows - 1)
    merged_ranges = list(load_list_output_sheet.merged_cells.ranges)

    for merged_range in merged_ranges:
        try:
            load_list_output_sheet.unmerge_cells(str(merged_range))
        except KeyError:
            continue

    for row in range(
        dynamic_start_row_number, template_row_number + total_rows
    ):  # Rows 4 to (4 + total_rows)
        # Apply the row height for the current row
        load_list_output_sheet.row_dimensions[row].height = template_row_height

        # Iterate through columns A to Q (1 to 17)
        for col in range(template_range_start_col, template_range_end_col + 1):
            # Get the template cell
            template_cell = load_list_output_sheet.cell(
                row=template_row_number, column=col
            )
            # Get the target cell
            target_cell = load_list_output_sheet.cell(row=row, column=col)
            # Copy the style from the template cell
            target_cell._style = template_cell._style

            # Apply column width (only once per column)
            column_letter = template_cell.column_letter
            if (
                row == dynamic_start_row_number
            ):  # Apply width only on the first iteration for each column
                template_width = load_list_output_sheet.column_dimensions[
                    column_letter
                ].width
                load_list_output_sheet.column_dimensions[column_letter].width = (
                    template_width
                )

    for index, data in enumerate(electrical_load_list_data):
        row = template_row_number + index
        load_list_output_sheet.cell(row=row, column=1, value=index + 1)
        load_list_output_sheet.cell(row=row, column=2, value=data.get("tag_number"))
        load_list_output_sheet.cell(
            row=row, column=3, value=data.get("service_description")
        )
        load_list_output_sheet.cell(row=row, column=4, value=data.get("working_kw"))
        load_list_output_sheet.cell(row=row, column=5, value=data.get("standby_kw"))
        load_list_output_sheet.cell(row=row, column=6, value=data.get("starter_type"))
        load_list_output_sheet.cell(
            row=row, column=7, value=f"{data.get('supply_voltage')} VAC"
        )
        load_list_output_sheet.cell(row=row, column=8, value=data.get("phase"))
        load_list_output_sheet.cell(
            row=row, column=9, value=data.get("motor_rated_current")
        )
        load_list_output_sheet.cell(
            row=row, column=10, value=data.get("control_scheme")
        )
        load_list_output_sheet.cell(row=row, column=11, value=data.get("panel"))
        load_list_output_sheet.cell(
            row=row, column=12, value=data.get("bus_segregation")
        )
        load_list_output_sheet.cell(
            row=row, column=13, value=data.get("motor_efficiency")
        )
        load_list_output_sheet.cell(row=row, column=14, value=data.get("package"))
        load_list_output_sheet.cell(row=row, column=15, value=data.get("area"))
        load_list_output_sheet.cell(row=row, column=16, value=data.get("remark"))
        load_list_output_sheet.cell(row=row, column=17, value=data.get("rev"))

    load_list_output_sheet.merge_cells("A1:Q1")
    # static row numbers which are calculation based
    calculated_row_start_number = total_rows + template_row_number

    # OUTGOING FEEDERS
    load_list_output_sheet.merge_cells(
        f"A{calculated_row_start_number}:Q{calculated_row_start_number}"
    )
    load_list_output_sheet.row_dimensions[calculated_row_start_number].height = 15

    # TOTAL POWER CONSUMPTION (Excluding Standby)
    load_list_output_sheet.merge_cells(
        f"A{calculated_row_start_number + 1}:C{calculated_row_start_number + 1}"
    )
    load_list_output_sheet.merge_cells(
        f"E{calculated_row_start_number + 1}:Q{calculated_row_start_number + 1}"
    )
    load_list_output_sheet.row_dimensions[calculated_row_start_number + 1].height = 30
    total_working_kw = sum(
        item.get("working_kw", 0) or 0 for item in electrical_load_list_data
    )
    load_list_output_sheet[f"D{calculated_row_start_number + 1}"] = total_working_kw

    # Row Gap
    load_list_output_sheet.merge_cells(
        f"A{calculated_row_start_number + 2}:Q{calculated_row_start_number + 2}"
    )
    load_list_output_sheet.row_dimensions[calculated_row_start_number + 2].height = 15

    # TOTAL POWER CONSUMPTION (Standby)
    load_list_output_sheet.merge_cells(
        f"A{calculated_row_start_number + 3}:C{calculated_row_start_number + 3}"
    )
    load_list_output_sheet.merge_cells(
        f"E{calculated_row_start_number + 3}:Q{calculated_row_start_number + 3}"
    )
    load_list_output_sheet.row_dimensions[calculated_row_start_number + 3].height = 30
    total_standby_kw = sum(
        item.get("standby_kw", 0) or 0 for item in electrical_load_list_data
    )
    load_list_output_sheet[f"D{calculated_row_start_number + 3}"] = total_standby_kw

    # Row Gap
    load_list_output_sheet.merge_cells(
        f"A{calculated_row_start_number + 4}:Q{calculated_row_start_number + 4}"
    )
    load_list_output_sheet.row_dimensions[calculated_row_start_number + 4].height = 15

    # TOTAL CONNECTED LOAD
    load_list_output_sheet.merge_cells(
        f"A{calculated_row_start_number + 5}:C{calculated_row_start_number + 5}"
    )
    load_list_output_sheet.merge_cells(
        f"E{calculated_row_start_number + 5}:Q{calculated_row_start_number + 5}"
    )
    load_list_output_sheet.row_dimensions[calculated_row_start_number + 5].height = 30
    load_list_output_sheet[f"D{calculated_row_start_number + 5}"] = total_working_kw

    # Row Gap
    load_list_output_sheet.merge_cells(
        f"A{calculated_row_start_number + 6}:Q{calculated_row_start_number + 6}"
    )
    load_list_output_sheet.row_dimensions[calculated_row_start_number + 6].height = 15

    # TOTAL LOAD
    load_list_output_sheet.merge_cells(
        f"A{calculated_row_start_number + 7}:C{calculated_row_start_number + 7}"
    )
    load_list_output_sheet.merge_cells(
        f"F{calculated_row_start_number + 7}:J{calculated_row_start_number + 7}"
    )
    load_list_output_sheet.merge_cells(
        f"K{calculated_row_start_number + 7}:Q{calculated_row_start_number + 7}"
    )
    if electrical_load_list_data and len(electrical_load_list_data) > 0:
        supply_voltage = electrical_load_list_data[0].get("supply_voltage", 0)
    else:
        supply_voltage = 0
    load_list_output_sheet.row_dimensions[calculated_row_start_number + 7].height = 30
    if supply_voltage == 0:
        total_load = 0
    else:
        total_load = total_working_kw * 1000 / (1.732 * supply_voltage * 0.8)

    load_list_output_sheet[f"D{calculated_row_start_number + 7}"] = round(total_load, 2)

    incomer_rating = total_load * 1.2 - 5
    first_higher_incomer = next(
        (
            incomer.get("incomer_rating", 0)
            for incomer in incomer_db_data
            if incomer["incomer_rating"] > incomer_rating
        ),
        incomer_rating,
    )
    first_higher_incomer = round(first_higher_incomer, 2)

    sheet_title = load_list_output_sheet.title

    if number_of_panels > 1 and sheet_title == "LOAD LIST OUTPUT Copy":
        load_list_output_sheet.delete_rows(calculated_row_start_number + 8, 2)
    else:
        # Row Gap
        load_list_output_sheet.merge_cells(
            f"A{calculated_row_start_number + 8}:Q{calculated_row_start_number + 8}"
        )
        load_list_output_sheet.row_dimensions[
            calculated_row_start_number + 8
        ].height = 15
        load_list_output_sheet[calculated_row_start_number + 8][0].alignment = (
            Alignment(horizontal="left", vertical="center")
        )

        load_list_output_sheet[f"B{calculated_row_start_number + 9}"] = panel_name
        load_list_output_sheet[f"C{calculated_row_start_number + 9}"] = (
            f"POWER SUPPLY \n {incomer_power_supply}"
        )
        load_list_output_sheet.merge_cells(
            f"D{calculated_row_start_number + 9}:E{calculated_row_start_number + 9}"
        )
        load_list_output_sheet[f"D{calculated_row_start_number + 9}"] = (
            f"I/C {first_higher_incomer} AMP \n MCCB"
        )

    return load_list_output_sheet
